from pathlib import Path
from openpyxl import load_workbook


def load_excel_emails(filename):

    file_path = Path("input") / filename

    if not file_path.exists():
        return []

    workbook = load_workbook(file_path)

    sheet = workbook.active

    emails = []

    for row in sheet.iter_rows(values_only=True):

        for value in row:

            if value:

                emails.append(str(value).strip())

    return emails